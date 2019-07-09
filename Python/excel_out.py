#!/usr/env python
# -*- coding:utf8 -*-
__author__ = '李国征'
__email__ = "745292907@qq.com"

from openpyxl import load_workbook, Workbook

wb = load_workbook("100000000061083.xlsx", read_only=True)

first_title = wb.get_sheet_by_name("first_title")
# c =first_title.cell(row=2, column=13)
# d =first_title.cell(row=3, column=13)
# e =first_title.cell(row=4, column=13)
# print(c.value, c.row)
# print(d.value, d.row)
# print(e.value, e.row)

# with open("/tmp/mi.txt", 'a+') as f:
#     for i in range(2,850002):
#         _tmp = first_title.cell(row=i, column=13)
#         e_id = _tmp.row
#         e_v = _tmp.value
#         f.write(e_v)

import time, random
from multiprocessing import JoinableQueue, Process


# def shengchanzhe(q, start, end):
#     for i in range(start, end):
#         _tmp = first_title.cell(row=i, column=13)
#         e_id = _tmp.row
#         e_v = _tmp.value
#         q.put("{},{}\n".format(e_id, e_v))
#
#     q.join()
#
# def xiaofeizhe(q):
#     while True:
#         food = q.get()
#         if food == None: break
#         print(food)
#
#         q.task_done()


# if __name__ == '__main__':
#     jq = JoinableQueue()
#     p1 = Process(target=shengchanzhe, args=(jq,2,100))
#     p1.start()
#     p2 = Process(target=shengchanzhe, args=(jq,100, 200))
#     p2.start()
#     p3 = Process(target=shengchanzhe, args=(jq,200, 300))
#     p3.start()
#
#     # c1 = Process(target=xiaofeizhe, args=(jq,))
#     # c1.daemon = True
#     # c1.start()
#     # c2 = Process(target=xiaofeizhe, args=(jq,))
#     # c2.daemon = True
#     # c2.start()
#     # c3 = Process(target=xiaofeizhe, args=(jq,))
#     # c3.daemon = True
#     # c3.start()
#     # c4 = Process(target=xiaofeizhe, args=(jq,))
#     # c4.daemon = True
#     # c4.start()
#     for p in range(100):
#         p = Process(target=xiaofeizhe, args=(jq,))
#         p.daemon = True
#         p.start()
#
#     p1.join()
#     p2.join()
#     p3.join()

from concurrent import futures
from threading import Thread, Semaphore, currentThread
sm = Semaphore(10)

def func(i):
    with sm:
        _tmp = first_title.cell(row=i, column=13)
        c = "{},{}\n".format(_tmp.row, _tmp.value)
        print(c)


# ft = futures.ThreadPoolExecutor(25)
# for i in range(2,10000):
#     t = ft.submit(func, i)
#     t.result()
for i in range(2, 10000):
    Thread(target=func, args=(i,)).start()


